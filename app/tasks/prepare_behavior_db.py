import os
import sqlite3
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

from app.core.config import resolve_path
from app.core.runtime import merge_runtime_status


EXPECTED_COLUMNS = [
    "tourist_id",
    "user_nickname",
    "age",
    "gender",
    "attraction_name",
    "attraction_content",
    "attraction_type",
    "visit_date",
    "stay_duration",
    "ticket_cost",
    "food_cost",
    "shopping_cost",
    "transport_cost",
    "entertainment_cost",
    "total_cost",
    "group_size",
    "satisfaction",
]


def normalize_header(value: object) -> str:
    return str(value).replace(" ", "").replace("\n", "").strip()


def find_behavior_excel(raw_dir: str) -> Path:
    candidates = sorted(Path(raw_dir).glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No behavior-analysis Excel files were found under: {raw_dir}")
    return candidates[0]


def load_behavior_rows(excel_path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    raw_headers = next(rows)
    headers = [normalize_header(value) for value in raw_headers]

    if not all(column in headers for column in EXPECTED_COLUMNS):
        raise RuntimeError(f"Unexpected behavior Excel headers: {headers}")

    records: list[dict[str, object]] = []
    for row in rows:
        if row is None or not any(value is not None and str(value).strip() for value in row):
            continue
        padded = list(row) + [None] * (len(headers) - len(row))
        record = dict(zip(headers, padded[: len(headers)]))
        records.append(record)
    return headers, records


def write_rows_to_sqlite(db_path: str, table_name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    column_defs = ", ".join(f'"{column}" TEXT' for column in headers)
    placeholders = ", ".join("?" for _ in headers)
    quoted_columns = ", ".join(f'"{column}"' for column in headers)
    insert_sql = f'INSERT INTO "{table_name}" ({quoted_columns}) VALUES ({placeholders})'

    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        cursor.execute(f'CREATE TABLE "{table_name}" ({column_defs})')
        cursor.executemany(insert_sql, [[row.get(column) for column in headers] for row in rows])
        conn.commit()


def prepare_behavior_db() -> Dict[str, object]:
    raw_dir = resolve_path("data/raw_sql_data")
    excel_path = find_behavior_excel(raw_dir)
    db_path = resolve_path("data/processed/tourist_behavior.db")

    headers, rows = load_behavior_rows(excel_path)

    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    write_rows_to_sqlite(db_path, "tourist_behavior", headers, rows)

    merge_runtime_status({"behavior_db_ready": os.path.exists(db_path)})
    return {
        "ok": True,
        "table": "tourist_behavior",
        "rows": len(rows),
        "db_path": db_path,
        "source_excel": str(excel_path),
    }
